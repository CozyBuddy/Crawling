import selenium
from selenium import webdriver
import time
from selenium.webdriver import ActionChains ,Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import bs4
import re
from datetime import datetime
import random
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import tempfile
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities


temp_profile_dir = tempfile.mkdtemp() # <- solution

opt = webdriver.ChromeOptions()
opt.set_capability("pageLoadStrategy", "normal")
# opt.add_argument(f"--user-data-dir={temp_profile_dir}")
#opt.add_argument('--headless')
opt.add_argument("--disable-extensions")
opt.add_argument("--disable-application-cache")
opt.add_experimental_option("prefs", {
    "profile.managed_default_content_settings.images": 2,
    "profile.managed_default_content_settings.plugins": 2,
    "profile.default_content_setting_values.notifications": 2 
})

opt.add_argument("--disable-popup-blocking")  # 팝업 차단
opt.add_argument("--disable-notifications")  # 알림 완전 차단

# opt.add_argument(f"--user-data-dir=/tmp/chrome_data_{debug_port}") 
# opt.add_argument(f"--remote-debugging-port={debug_port}")

opt.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                 "AppleWebKit/537.36 (KHTML, like Gecko) "
                 "Chrome/119.0.6045.200 Safari/537.36")

opt.add_argument("--disable-blink-features=AutomationControlled"); 
opt.add_argument("--disable-dev-shm-usage")
opt.add_argument("--disable-gpu")
browser = webdriver.Chrome(options=opt )
browser.set_script_timeout(3)
browser.set_page_load_timeout(5)
browser.implicitly_wait(10)
wait = WebDriverWait(browser, 10)
num = 1
wb = Workbook()
ws = wb.active
excelnum = 1
while True :
    if num >= 335 :
        break
    try:
        browser.get(f'https://www.hani.co.kr/arti/politics?page={num}')

        #results = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.BaseArticleCard_content__tYkEA > a[href]")))
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )

        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            #result3 = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.article-text')))
            
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break

    num+=1

wb.save(f'뉴스기사수집(한겨레){datetime.now().date()}.xlsx')
browser.quit()
browser = webdriver.Chrome(options=opt)
wb2 = Workbook()
ws2 = wb2.active
excelnum = 1
num=1
while True :
    if num >= 334 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/society?page={num}')
        
        # results = wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div.BaseArticleCard_content__tYkEA > a[href]")))
        # #time.sleep(3)
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )
        #results2 = browser.find_elements(By.CSS_SELECTOR , 'div.section_article._TEMPLATE div.sa_text > a' )
        #results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )
        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            #time.sleep(3)

            #result3 = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.article-text')))
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1

wb2.save(f'뉴스기사수집(한겨레)2{datetime.now().date()}.xlsx')

browser.quit()
browser = webdriver.Chrome(options=opt)
wb3 = Workbook()
ws3 = wb3.active
num=1
excelnum = 1
while True :
    if num >= 334 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/area?page={num}')
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )

        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1

wb3.save(f'뉴스기사수집(한겨레)3{datetime.now().date()}.xlsx')
browser.quit()
browser = webdriver.Chrome(options=opt)
excelnum = 1
wb4 = Workbook()
ws4 = wb4.active
num=1
while True :
    if num >= 334 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/economy?page={num}')
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )
        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1

wb4.save(f'뉴스기사수집(한겨레)4{datetime.now().date()}.xlsx')
excelnum = 1
wb5 = Workbook()
ws5 = wb5.active
num=1
while True :
    if num >= 334 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/international?page={num}')
        
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )
        
        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1

wb5.save(f'뉴스기사수집(한겨레)5{datetime.now().date()}.xlsx')

browser.quit()
browser = webdriver.Chrome(options=opt)
excelnum = 1
num=1
while True :
    if num >= 334 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/culture?page={num}')
        
        #results = wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div.BaseArticleCardVertical_card__LLa8l > a[href]")))
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCardVertical_card__LLa8l > a' )
        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
       
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1
num=1

excelnum = 1

browser.quit()
browser = webdriver.Chrome(options=opt)
wb6 = Workbook()
ws6 = wb6.active
while True :
    if num >= 335 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/science?page={num}')

        #results = wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div.BaseArticleCard_content__tYkEA > a[href]")))
        #results2 = browser.find_elements(By.CSS_SELECTOR , 'div.section_article._TEMPLATE div.sa_text > a' )
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )
        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1

wb6.save(f'뉴스기사수집(한겨레)6{datetime.now().date()}.xlsx')

excelnum = 1

browser.quit()
browser = webdriver.Chrome(options=opt)
wb7 = Workbook()
ws7 = wb7.active
num=1
while True :
    if num >= 335 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/animalpeople?page={num}')
        
        
        # results = wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "a.ArticleListTitle_link__U_6Bo[href]")))
        #results2 = browser.find_elements(By.CSS_SELECTOR , 'div.section_article._TEMPLATE div.sa_text > a' )
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'a.ArticleListTitle_link__U_6Bo' )
        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
         
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1

wb7.save(f'뉴스기사수집(한겨레)7{datetime.now().date()}.xlsx')
excelnum = 1

browser.quit()
browser = webdriver.Chrome(options=opt)
wb8 = Workbook()
ws8 = wb8.active
num=1
while True :
    if num >= 335 :
        break
    try:

        browser.get(f'https://www.hani.co.kr/arti/opinion?page={num}')
        
        #results = wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div.BaseArticleCard_content__tYkEA > a[href]")))
        #results2 = browser.find_elements(By.CSS_SELECTOR , 'div.section_article._TEMPLATE div.sa_text > a' )
        time.sleep(3)
        results = browser.find_elements(By.CSS_SELECTOR , 'div.BaseArticleCard_content__tYkEA > a' )
        href_list = [el.get_attribute('href') for el in results]
    except Exception as e :
        num+=1
        print('에러' ,e)
        continue
   # href_list2 = [el.get_attribute('href') for el in results2]
    for a in href_list:
        try:
            browser.get(a)
            
            time.sleep(1.5)
            result3 = browser.find_element(By.CSS_SELECTOR , 'div.article-text' )
            splitresult = [s.strip() for s in result3.text.split('.') if s.strip() ]
            #.replace('=','').replace('@','')
            for i in splitresult:
               # if re.fullmatch(r"[가-힣0-9.·,!=?\"'()\-\ ]+", i) :
                ws[f"A{excelnum+1}"] = i 
                excelnum+=1
        
        except Exception as e:
            print(f"Error at {e}")
            continue
    if num >= 334 :
        break
    num+=1

wb8.save(f'뉴스기사수집(한겨레)8{datetime.now().date()}.xlsx')


