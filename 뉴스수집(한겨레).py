import asyncio
from playwright.async_api import async_playwright
from openpyxl import Workbook
from datetime import datetime
import re

async def main():
    wb = Workbook()
    ws = wb.active

    wb2 = Workbook()
    ws2 = wb2.active

    wb3 = Workbook()
    ws3 = wb3.active

    wb4 = Workbook()
    ws4 = wb4.active

    wb5 = Workbook()
    ws5 = wb5.active

    wb6 = Workbook()
    ws6 = wb6.active

    wb7 = Workbook()
    ws7 = wb7.active

    wb8 = Workbook()
    ws8 = wb8.active

    wb9 = Workbook()
    ws9 = wb9.active

    excelnum = 1
    num = 1

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=[
            "--disable-gpu",
            "--disable-blink-features=AutomationControlled",
            "--disable-extensions",
            "--disable-dev-shm-usage",
            "--no-sandbox"
        ])
        context = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/119.0.6045.200 Safari/537.36"),
            permissions=[]
        )

        page = await context.new_page()

        while num < 335:
            try:
                url = f'https://www.hani.co.kr/arti/politics?page={num}'
                await page.goto(url, timeout=15000)

                # 링크들 수집
                href_list = await page.eval_on_selector_all(
                    "div.BaseArticleCard_content__tYkEA > a",
                    "els => els.map(e => e.href)"
                )

                for a in href_list:
                    try:
                        await page.goto(a, timeout=15000)
                        await page.wait_for_selector("div.article-text", timeout=8000)
                        text = await page.inner_text("div.article-text")
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e:
                        print(f"Error in article: {e}")
                        continue

            except Exception as e:
                print(f"Error in page {num}: {e}")
                continue
            num += 1

        # await browser.close()

        wb.save(f'뉴스기사수집(한겨레){datetime.now().date()}.xlsx')
        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/society?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.BaseArticleCard_content__tYkEA > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws2[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb2.save(f'뉴스기사수집(한겨레)2{datetime.now().date()}.xlsx')   

        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/area?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.BaseArticleCard_content__tYkEA > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws3[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb3.save(f'뉴스기사수집(한겨레)3{datetime.now().date()}.xlsx') 

        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/economy?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.BaseArticleCard_content__tYkEA > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws4[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb4.save(f'뉴스기사수집(한겨레)4{datetime.now().date()}.xlsx') 


        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/international?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.BaseArticleCard_content__tYkEA > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws5[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb5.save(f'뉴스기사수집(한겨레)5{datetime.now().date()}.xlsx') 

        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/culture?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.BaseArticleCardVertical_card__LLa8l > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws6[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb6.save(f'뉴스기사수집(한겨레)6{datetime.now().date()}.xlsx') 

        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/science?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.BaseArticleCard_content__tYkEA > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws7[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb7.save(f'뉴스기사수집(한겨레)7{datetime.now().date()}.xlsx') 

        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/animalpeople?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.ArticleListTitle_link__U_6Bo > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws8[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb8.save(f'뉴스기사수집(한겨레)8{datetime.now().date()}.xlsx') 

        num =1
        excelnum=1
        while num < 355 :
            try:
                url = f'https://www.hani.co.kr/arti/opinion?page={num}'
                await page.goto(url,timeout=8000)

                href_list = await page.eval_on_selector_all(
                    'div.BaseArticleCard_content__tYkEA > a',
                    'els => els.map(e=> e.href)'
                )

                for a in href_list:
                    try:
                        await page.goto(a , timeout=8000)
                        await page.wait_for_selector('div.article-text' , timeout=8000)

                        text = page.inner_text('div.article-text')
                        splitresult = [s.strip() for s in text.split('.') if s.strip()]

                        for i in splitresult:
                            ws9[f"A{excelnum+1}"] = i
                            excelnum += 1
                    except Exception as e :
                        print(f"Error in article: {e}")
                        continue
            except Exception as e :
                print(f"Error in page {num}: {e}")
                continue
        wb9.save(f'뉴스기사수집(한겨레)9{datetime.now().date()}.xlsx') 

asyncio.run(main())
