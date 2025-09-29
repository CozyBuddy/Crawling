import requests
import bs4 
import urllib.request as req
import os
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage


book = openpyxl.Workbook()
sheet = book.active

code = requests.get('http://books.toscrape.com/')

soup = bs4.BeautifulSoup(code.text , 'html.parser')

result = soup.select('div.image_container img')
print(result)
result2 = soup.select('h3 > a')
result3 = soup.select('div.product_price > p.price_color')

if not os.path.exists('./CGV'):
        os.mkdir('./CGV')

num = 1
for x, y ,z in zip(result, result2 ,result3):
    print(f'제목:{y.text}')
    print(z.text)

    url = 'http://books.toscrape.com/'+x.attrs['src']

    req.urlretrieve(url, f"CGV/{y.text}.jpg")

    img = PILImage.open(f"CGV/{y.text}.jpg")
    img_resized = img.resize((int(img.width*0.9) , int(img.height*0.9)))
    img_resized.save(f"CGV/{y.text}_resized.jpg")
    img.close()
    sheet.add_image(Image(f"CGV/{y.text}_resized.jpg"), f'A{num}')
    sheet.cell(row=num , column=2).value = y.text
    sheet.cell(row=num , column=3).value = z.text

    sheet.row_dimensions[num].height = 197
    num +=1
    print('-'*100)
    
sheet.column_dimensions['A'].width = 23
sheet.column_dimensions['B'].width = 35
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 21
book.save('./CGV_무비차트.xlsx')
